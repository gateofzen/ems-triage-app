# v20 - Excel テンプレート直接書込み方式（座標ズレ完全解消）
import streamlit as st
import cv2
import numpy as np
from PIL import Image
import io
import base64
import openpyxl
from pyzbar.pyzbar import decode
import os
import textwrap
from datetime import datetime

def decode_qr_raw(b64_string):
    return base64.b64decode(b64_string).decode("utf-8").split(",")

def parse_ems_qr(items):
    try:
        dt = datetime.strptime(items[1], "%Y/%m/%d %H:%M:%S")
        weeks = ["月","火","水","木","金","土","日"]
        name_raw = items[4]
        kanji, kana = (name_raw.split("θ",1) if "θ" in name_raw else (name_raw,""))
        field_8 = items[8].strip() if len(items)>8 else ""
        field_9 = items[9].strip() if len(items)>9 else ""
        field_10 = items[10].strip() if len(items)>10 else ""
        field_11 = items[11].strip() if len(items)>11 else ""
        complaint = ""
        history = field_9 if len(field_9) > len(field_8) else field_8
        for f in [field_8, field_10, field_11]:
            if f and len(f) <= 30 and not f.isdigit():
                complaint = f; break
        if not complaint and history:
            for sep in ["。","、"]:
                if sep in history:
                    c = history[:history.index(sep)]
                    if len(c) <= 30: complaint = c; break
        return {
            "month": str(dt.month), "day": str(dt.day),
            "weekday": weeks[dt.weekday()],
            "hour": f"{dt.hour:02}", "minute": f"{dt.minute:02}",
            "kanji": kanji.strip().replace("　"," "),
            "kana": kana.strip().replace("　"," "),
            "gender": items[5],
            "complaint": complaint, "history": history,
            "birth": items[13],
            "age": items[14] if len(items)>14 else "",
            "jcs": items[15] if len(items)>15 else "",
            "bp_s": items[19] if len(items)>19 else "",
            "bp_d": items[20] if len(items)>20 else "",
            "hr": items[21] if len(items)>21 else "",
            "rr": items[22] if len(items)>22 else "",
            "bt": items[23] if len(items)>23 else "",
            "spo2": items[24] if len(items)>24 else "",
        }
    except Exception as e:
        st.error(f"データ解析失敗: {e}")
        return None

st.set_page_config(page_title="台帳作成システム", layout="centered")
st.title("🚑 トリアージ台帳 v20")

uploaded_file = st.file_uploader("QRコードのスクリーンショットを選択", type=["png","jpg","jpeg"])

if uploaded_file:
    img = Image.open(uploaded_file)
    with st.spinner("QRコードをスキャン中..."):
        qr = decode(img)
        if not qr:
            qr = decode(cv2.cvtColor(np.array(img), cv2.COLOR_RGB2GRAY))
        if not qr:
            qr = decode(cv2.detailEnhance(np.array(img), sigma_s=10, sigma_r=0.15))

    if qr:
        items = decode_qr_raw(qr[0].data.decode("utf-8"))
        data = parse_ems_qr(items)
        if data:
            st.success(f"読込成功: {data['kanji']} 様")

            with st.expander("🔍 QRデータ確認"):
                labels = {1:"依頼日時",4:"氏名",5:"性別",8:"主訴",9:"経過等",
                          13:"生年月日",14:"年齢",15:"JCS",19:"BP上",20:"BP下",
                          21:"HR",22:"RR",23:"BT",24:"SpO2"}
                for i, v in enumerate(items):
                    lbl = f"  ← {labels[i]}" if i in labels else ""
                    st.text(f"[{i:2d}] {v[:80]}{lbl}")

            st.subheader("台帳情報の入力")
            recorder = st.selectbox("記載者", ["前川","森木","小舘","遠藤"])
            col1, col2 = st.columns(2)
            with col1:
                origin = st.text_input("依頼元（救急隊）", value="中央").replace("救急隊","").strip()
                history_yn = st.radio("受診歴", ["無","有"], index=0, horizontal=True)
            with col2:
                history_dept = st.text_input("受診科名（有の場合）")
                decision = st.radio("判定", ["応需","不応需"], index=0, horizontal=True)
            complaint_edit = st.text_input("主訴（編集可）", value=data.get("complaint",""))

            res = {}
            if decision == "応需":
                res["init"] = st.selectbox("初期対応した科", ["当直医","救急科","その他"])
                if res["init"] == "その他":
                    res["init_other"] = st.text_input("初期対応科名")
                res["out"] = st.selectbox("最終転帰", ["入院","帰宅","その他"])
                if res["out"] == "入院":
                    res["ward"] = st.selectbox("病棟", ["4東","HCU","ICU","その他"])
                    if res["ward"] == "その他":
                        res["ward_other"] = st.text_input("病棟名")
                    res["main"] = st.selectbox("主科", ["臨研","救急科","その他"])
                    if res["main"] == "その他":
                        res["main_other"] = st.text_input("主科名")
            else:
                reasons = ["1. 緊急性なし","2. ベッド満床","3. 既定の応需不可症例",
                           "4. 対応可能な医師不在","5. 緊急手術制限中",
                           "6-A. 医師処置中につき対応困難","6-B. 看護師処置中につき対応困難","7. その他"]
                res["reason"] = st.selectbox("不応需理由", reasons)

            if st.button("台帳を生成する", type="primary"):
                with st.spinner("作成中..."):
                    wb = openpyxl.load_workbook("トリアージ台帳.xlsx")
                    ws = wb.active

                    # ===== 記載者 =====
                    ws["H2"] = recorder

                    # ===== 依頼日時 =====
                    date_str = f"{data['month']}/{data['day']}  ({data['weekday']})  {data['hour']}:{data['minute']}"
                    ws["B3"] = date_str

                    # ===== 依頼元 =====
                    ws["F3"] = f"{origin}救急隊"

                    # ===== 受診歴 =====
                    if history_yn == "有":
                        dept = history_dept if history_dept else ""
                        ws["I3"] = f"◯有( {dept} 科)・無"
                    else:
                        ws["I3"] = "有(          科)・◯無"

                    # ===== 患者名 =====
                    ws["B6"] = f"{data['kana']}\n{data['kanji']}"

                    # ===== 生年月日 =====
                    b = data["birth"]
                    if len(b) >= 8:
                        ws["H6"] = f"{b[:4]}年"
                        ws["I6"] = f"{b[4:6]}月"
                        ws["J6"] = f"{b[6:8]}日"

                    # ===== 年齢 =====
                    if data["age"]:
                        ws["G8"] = f"{data['age']}歳"

                    # ===== 性別 =====
                    if data["gender"] == "2" or "女" in data["gender"]:
                        ws["I8"] = "男　・　◯女"
                    else:
                        ws["I8"] = "◯男　・　女"

                    # ===== 主訴 =====
                    ws["B10"] = complaint_edit

                    # ===== 経過等 =====
                    ws["B14"] = data["history"]

                    # ===== バイタルサイン =====
                    if data["jcs"]:
                        ws["I20"] = f"JCS {data['jcs']}"
                    if data["rr"]:
                        ws["I22"] = f"{data['rr']} 回/分"
                    if data["hr"]:
                        ws["I24"] = f"{data['hr']} 回/分"
                    if data["bp_s"] and data["bp_d"]:
                        ws["I26"] = f"{data['bp_s']}/{data['bp_d']} mmHg"
                    if data["spo2"]:
                        ws["I28"] = f"{data['spo2']} %"
                    if data["bt"]:
                        ws["I30"] = f"{data['bt']} ℃"

                    # ===== 判定 =====
                    if decision == "応需":
                        ws["A38"] = "◯応需\n・\n不応需\n(どちらかに◯)"

                        # 初期対応した科
                        if res["init"] == "当直医":
                            ws["D38"] = "◯当直医　　・救急科　　・その他(　　　　　　　　科)"
                        elif res["init"] == "救急科":
                            ws["D38"] = "・当直医　　◯救急科　　・その他(　　　　　　　　科)"
                        elif res["init"] == "その他":
                            other_name = res.get("init_other","")
                            ws["D38"] = f"・当直医　　・救急科　　◯その他( {other_name} )"

                        # 最終転帰
                        if res["out"] == "入院":
                            ws["C41"] = "◯入院\n・帰宅\nその他(　　　　　)"

                            # 病棟
                            ward = res.get("ward","")
                            if ward == "4東":
                                ws["F42"] = "◯4東　・HCU　・ICU\n・　　　　　　　　　病棟"
                            elif ward == "HCU":
                                ws["F42"] = "・4東　◯HCU　・ICU\n・　　　　　　　　　病棟"
                            elif ward == "ICU":
                                ws["F42"] = "・4東　・HCU　◯ICU\n・　　　　　　　　　病棟"
                            elif ward == "その他":
                                ward_name = res.get("ward_other","")
                                ws["F42"] = f"・4東　・HCU　・ICU\n・　{ward_name}　病棟"

                            # 主科
                            main = res.get("main","")
                            if main == "臨研":
                                ws["I42"] = "◯臨研"
                            elif main == "救急科":
                                ws["I43"] = "◯救急科"
                            elif main == "その他":
                                main_name = res.get("main_other","").rstrip("科")
                                ws["I44"] = f"・( {main_name} )科"

                        elif res["out"] == "帰宅":
                            ws["C41"] = "・入院\n◯帰宅\nその他(　　　　　)"
                    else:
                        ws["A38"] = "応需\n・\n◯不応需\n(どちらかに◯)"

                        # 不応需理由にマーク
                        reason_cells = {"1":"B45","2":"B46","3":"B47","4":"B48",
                                       "5":"B49","6-A":"B50","6-B":"B51","7":"B52"}
                        rk = res["reason"].split(".")[0].strip()
                        if rk in reason_cells:
                            cell = reason_cells[rk]
                            original = ws[cell].value or ""
                            ws[cell] = "◯" + original

                    # ===== Excelファイルとして保存 =====
                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)

                    st.success("台帳を生成しました")
                    st.download_button(
                        "📥 台帳をダウンロード（Excel）",
                        buf.getvalue(),
                        f"triage_{data['kanji']}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
    else:
        st.error("QRコードが見つかりません。画像を拡大するか、明るい場所で撮り直してください。")
